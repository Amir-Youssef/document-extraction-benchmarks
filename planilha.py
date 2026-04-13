"""
gerar_planilha.py
=================
Lê uma pasta com subpastas contendo arquivos JSON de métricas de extração
e gera uma planilha Excel com 4 abas no padrão estabelecido.

Estrutura esperada:
    resultados/
        results_modelo_a/
            metrics_tipo1_modelo_a_TIMESTAMP.json
            metrics_tipo2_modelo_a_TIMESTAMP.json
        results_modelo_b/
            metrics_tipo1_modelo_b_TIMESTAMP.json
            ...

Uso:
    python gerar_planilha.py                          # pasta padrão: ./resultados
    python gerar_planilha.py /caminho/para/resultados
    python gerar_planilha.py /caminho/para/resultados planilha_saida.xlsx
"""

import sys
import json
import glob
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ── Cores e labels ─────────────────────────────────────────────────────────────
# Adicione ou altere entradas aqui para personalizar cores por modelo.
# Qualquer modelo não listado receberá uma cor padrão automaticamente.
MODEL_COLORS_PRESET = {
    "gemini_3_flash"                 : "1565C0",
    "gemini_3_1_pro"                 : "006064",
    "kimi_k25"                     : "6A1B9A",
    "ollama_gemma3_12b"            : "F57F17",
    "ollama_gemma3_27b"            : "BF360C",
    "ollama_qwen3_vl"              : "1B5E20",
    "ollama_qwen3_vl_30b"          : "2E7D32",
    "ollama_qwen3_vl_8b_latest"    : "388E3C",
    "qwen3_vl"                     : "1B5E20",
    "qwen3_vl_8b"                  : "388E3C",
    "qwen3_vl_bb"                  : "388E3C",
    "qwen3.5_27b"                  : "BF360C",
    "qwen3.5_35b"                   : "E64A19",
    "qwen3.5_9b"                   : "F4511E",
}

MODEL_LABELS_PRESET = {
    "gemini_3_flash"                 : "Gemini 3 Flash",
    "gemini_3_1_pro"                 : "Gemini 3.1 Pro",
    "kimi_k25"                     : "Kimi K2.5",
    "ollama_gemma3_12b"            : "Gemma3 12B (Ollama)",
    "ollama_gemma3_27b"            : "Gemma3 27B (Ollama)",
    "ollama_qwen3_vl"              : "Qwen3-VL (Ollama)",
    "ollama_qwen3_vl_30b"          : "Qwen3-VL 30B (Ollama)",
    "ollama_qwen3_vl_8b_latest"    : "Qwen3-VL 8B Latest",
    "qwen3_vl"                     : "Qwen3-VL",
    "qwen3_vl_8b"                  : "Qwen3-VL 8B",
    "qwen3_vl_bb"                  : "Qwen3-VL 8B (v2)",
    "qwen3.5_27b"                  : "Qwen3.5 27B",
    "qwen3.5_35b"                  : "Qwen3.5 35B",
    "qwen3.5_9b"                   : "Qwen3.5 9B",
}

# Cores de fallback para modelos não mapeados (ciclo automático)
FALLBACK_COLORS = [
    "37474F", "4A148C", "1A237E", "004D40", "BF360C",
    "880E4F", "F57F17", "0D47A1", "33691E", "4E342E",
]

BASE_COLORS = {
    "title_bg"   : "1F3864",
    "subtitle_bg": "243F60",
    "alt1"       : "F5F8FF",
    "alt2"       : "FFFFFF",
    "sum_row"    : "37474F",
    "err_hdr"    : "B71C1C",
    "err_light"  : "FFEBEE",
    "err_alt"    : "FFF5F5",
}

PALETTE = [
    ("≥ 95%", "63BE7B", "1A4001"),
    ("85–95%", "A8D96A", "2D5A00"),
    ("70–85%", "FFD700", "7D6608"),
    ("55–70%", "FFA552", "7D2F00"),
    ("<  55%", "FF6B6B", "7D0000"),
]


# ── Helpers de estilo ──────────────────────────────────────────────────────────
def fnt(bold=False, size=9, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def fill(c):
    return PatternFill("solid", fgColor=c)

def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr(bottom_medium=False):
    t = Side(style="thin")
    b = Side(style="medium" if bottom_medium else "thin")
    return Border(left=t, right=t, top=t, bottom=b)

def cscale():
    return ColorScaleRule(
        start_type="num", start_value=0.7, start_color="FF6B6B",
        mid_type="num",   mid_value=0.9,   mid_color="FFD700",
        end_type="num",   end_value=1.0,   end_color="63BE7B",
    )

def make_banner(ws, title, subtitle, n_cols):
    ws.sheet_view.showGridLines = False
    lc = get_column_letter(n_cols)
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A1:{lc}1")
    c = ws["A1"]; c.value = title
    c.font = fnt(bold=True, size=14, color="FFFFFF")
    c.fill = fill(BASE_COLORS["title_bg"]); c.alignment = aln()
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{lc}2")
    c = ws["A2"]; c.value = subtitle
    c.font = fnt(size=10, color="BBBBBB", italic=True)
    c.fill = fill(BASE_COLORS["subtitle_bg"]); c.alignment = aln()
    ws.row_dimensions[3].height = 5
    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = fill(BASE_COLORS["title_bg"])


# ── Carregar dados ─────────────────────────────────────────────────────────────
def load_data(root: Path):
    all_data = {}
    for f in sorted(root.rglob("*.json")):
        try:
            with open(f, encoding="utf-8") as fp:
                d = json.load(fp)
            key = (d["tipo_documento"], d["modelo"])
            if key not in all_data or d["timestamp"] > all_data[key]["timestamp"]:
                all_data[key] = d
        except Exception as e:
            print(f"  [AVISO] Erro ao ler {f}: {e}")
    return all_data


# ── Montar labels e cores por modelo ──────────────────────────────────────────
def build_model_meta(modelos):
    labels = {}
    colors = {}
    fallback_idx = 0
    for m in modelos:
        labels[m] = MODEL_LABELS_PRESET.get(m, m.replace("_", " ").title())
        if m in MODEL_COLORS_PRESET:
            colors[m] = MODEL_COLORS_PRESET[m]
        else:
            colors[m] = FALLBACK_COLORS[fallback_idx % len(FALLBACK_COLORS)]
            fallback_idx += 1
    return labels, colors


# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — Visão Geral
# ══════════════════════════════════════════════════════════════════════════════
def build_visao_geral(wb, all_data, TIPOS, MODELOS, MODEL_LABELS, MODEL_COLORS, date_range):
    N = 4  # Accuracy | Levenshtein | Lat.(s) | Diverg.
    ws = wb.create_sheet("📊 Visão Geral")
    ws.freeze_panes = "B9"
    TC = 1 + len(MODELOS) * N
    modelos_str = " · ".join(MODEL_LABELS[m] for m in MODELOS)

    make_banner(ws,
        "COMPARAÇÃO DE MODELOS — RESULTADOS DE EXTRAÇÃO",
        f"{modelos_str}  ·  {date_range}",
        TC)

    ws.column_dimensions["A"].width = 24
    for col in range(2, TC + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Cabeçalhos modelo (linhas 7–8)
    ws.row_dimensions[7].height = 22; ws.row_dimensions[8].height = 26
    ws.merge_cells("A7:A8")
    ws["A7"].value = "Tipo de Documento"
    ws["A7"].font = fnt(bold=True, size=10, color="FFFFFF")
    ws["A7"].fill = fill(BASE_COLORS["title_bg"]); ws["A7"].alignment = aln()

    SUB = ["Accuracy", "Levenshtein", "Lat. (s)", "Diverg."]
    for mi, model in enumerate(MODELOS):
        cs = 2 + mi * N; ce = cs + N - 1
        ws.merge_cells(f"{get_column_letter(cs)}7:{get_column_letter(ce)}7")
        c = ws[f"{get_column_letter(cs)}7"]
        c.value = MODEL_LABELS[model]; c.font = fnt(bold=True, size=10, color="FFFFFF")
        c.fill = fill(MODEL_COLORS[model]); c.alignment = aln()
        for i, sub in enumerate(SUB):
            c = ws.cell(row=8, column=cs + i)
            c.value = sub; c.font = fnt(bold=True, size=9, color="FFFFFF")
            c.fill = fill(MODEL_COLORS[model]); c.alignment = aln(wrap=True)

    # Dados
    DS = 9
    for ri, tipo in enumerate(TIPOS):
        row = DS + ri; bg = BASE_COLORS["alt1"] if ri % 2 == 0 else BASE_COLORS["alt2"]
        ws.row_dimensions[row].height = 17
        c = ws.cell(row=row, column=1, value=tipo.upper().replace("_", " "))
        c.font = fnt(bold=True, size=9); c.fill = fill(bg); c.alignment = aln(h="left")
        for mi, model in enumerate(MODELOS):
            cs = 2 + mi * N; d = all_data.get((tipo, model))
            if not d:
                for i in range(N):
                    cell = ws.cell(row=row, column=cs + i, value="N/D")
                    cell.fill = fill(bg); cell.alignment = aln()
                    cell.font = fnt(size=9, color="AAAAAA")
                continue
            mg = d["metricas_gerais"]
            vals = [
                mg.get("mean_accuracy"),
                mg.get("mean_levenshtein"),
                round(mg.get("latencia", {}).get("mean", 0), 2),
                f"{mg.get('docs_com_divergencia', 0)}/{mg.get('docs_processados', 0)} ({mg.get('total_docs', 0)})",
            ]
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=cs + i)
                if isinstance(v, float) and i < 2:
                    cell.value = v; cell.number_format = "0.00%"
                elif isinstance(v, float):
                    cell.value = v; cell.number_format = "0.00"
                else:
                    cell.value = v
                cell.fill = fill(bg); cell.alignment = aln(); cell.font = fnt(size=9)

    # Média geral
    sr = DS + len(TIPOS)
    ws.row_dimensions[sr].height = 20
    c = ws.cell(row=sr, column=1, value="MÉDIA GERAL")
    c.font = fnt(bold=True, size=9, color="FFFFFF")
    c.fill = fill(BASE_COLORS["sum_row"]); c.alignment = aln(h="left")
    for mi in range(len(MODELOS)):
        cs = 2 + mi * N
        for i in range(2):
            col = get_column_letter(cs + i)
            cell = ws.cell(row=sr, column=cs + i,
                           value=f"=AVERAGE({col}{DS}:{col}{sr - 1})")
            cell.number_format = "0.00%"; cell.font = fnt(bold=True, color="FFFFFF")
            cell.fill = fill(BASE_COLORS["sum_row"]); cell.alignment = aln()
        col = get_column_letter(cs + 2)
        cell = ws.cell(row=sr, column=cs + 2,
                       value=f"=AVERAGE({col}{DS}:{col}{sr - 1})")
        cell.number_format = "0.00"; cell.font = fnt(bold=True, color="FFFFFF")
        cell.fill = fill(BASE_COLORS["sum_row"]); cell.alignment = aln()
        cell = ws.cell(row=sr, column=cs + 3, value="—")
        cell.font = fnt(bold=True, color="FFFFFF")
        cell.fill = fill(BASE_COLORS["sum_row"]); cell.alignment = aln()

    # Bordas e formatação condicional
    for row in range(7, sr + 1):
        for col in range(1, TC + 1):
            ws.cell(row=row, column=col).border = bdr()
    for mi in range(len(MODELOS)):
        cs = 2 + mi * N
        for i in range(2):
            col = get_column_letter(cs + i)
            ws.conditional_formatting.add(f"{col}{DS}:{col}{sr - 1}", cscale())


# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — Métricas por Campo
# ══════════════════════════════════════════════════════════════════════════════
def build_metricas_campo(wb, all_data, TIPOS, MODELOS, MODEL_LABELS, MODEL_COLORS):
    ws = wb.create_sheet("🔍 Métricas por Campo")
    ws.freeze_panes = "C6"
    TC2 = 2 + len(MODELOS) * 2
    make_banner(ws,
        f"MÉTRICAS POR CAMPO — {len(MODELOS)} MODELOS",
        f"Accuracy e Levenshtein por campo extraído — {len(TIPOS)} tipos de documento",
        TC2)

    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 24
    for col in range(3, TC2 + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 28
    for ref, val in [("A4", "Tipo"), ("B4", "Campo")]:
        c = ws[ref]; c.value = val
        c.font = fnt(bold=True, size=9, color="FFFFFF")
        c.fill = fill(BASE_COLORS["title_bg"]); c.alignment = aln()
    ws.merge_cells("A4:A5"); ws.merge_cells("B4:B5")

    for mi, model in enumerate(MODELOS):
        cs = 3 + mi * 2
        ws.merge_cells(f"{get_column_letter(cs)}4:{get_column_letter(cs + 1)}4")
        c = ws[f"{get_column_letter(cs)}4"]
        c.value = MODEL_LABELS[model]; c.font = fnt(bold=True, size=9, color="FFFFFF")
        c.fill = fill(MODEL_COLORS[model]); c.alignment = aln()
        for i, h in enumerate(["Accuracy", "Levenshtein"]):
            c = ws.cell(row=5, column=cs + i)
            c.value = h; c.font = fnt(bold=True, size=8, color="FFFFFF")
            c.fill = fill(MODEL_COLORS[model]); c.alignment = aln(wrap=True)

    row2 = 6
    for tipo in TIPOS:
        campos = sorted(set(
            campo for model in MODELOS
            for d in [all_data.get((tipo, model))] if d
            for campo in d.get("metricas_por_campo", {}).keys()
        ))
        for j, campo in enumerate(campos):
            bg = BASE_COLORS["alt1"] if j % 2 == 0 else BASE_COLORS["alt2"]
            ws.row_dimensions[row2].height = 15
            c = ws.cell(row=row2, column=1,
                        value=tipo.upper().replace("_", " ") if j == 0 else "")
            c.font = fnt(bold=(j == 0), size=9)
            c.fill = fill(bg); c.alignment = aln(h="left", v="top")
            c = ws.cell(row=row2, column=2, value=campo)
            c.font = fnt(size=9); c.fill = fill(bg); c.alignment = aln(h="left")
            for mi, model in enumerate(MODELOS):
                cs = 3 + mi * 2; d = all_data.get((tipo, model))
                mf = d["metricas_por_campo"].get(campo, {}) if d else {}
                for i, key in enumerate(["accuracy", "mean_levenshtein"]):
                    v = mf.get(key)
                    cell = ws.cell(row=row2, column=cs + i)
                    cell.value = round(v, 4) if v is not None else "—"
                    if v is not None: cell.number_format = "0.00%"
                    cell.fill = fill(bg); cell.alignment = aln(); cell.font = fnt(size=9)
            is_last = (j == len(campos) - 1)
            for col in range(1, TC2 + 1):
                ws.cell(row=row2, column=col).border = bdr(bottom_medium=is_last)
            row2 += 1

    for mi in range(len(MODELOS)):
        col = get_column_letter(3 + mi * 2)
        ws.conditional_formatting.add(f"{col}6:{col}{row2}", cscale())


# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — Divergências
# ══════════════════════════════════════════════════════════════════════════════
def build_divergencias(wb, all_data, TIPOS, MODELOS, MODEL_LABELS):
    ws = wb.create_sheet("⚠️ Divergências")
    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A5"
    for col, w in zip("ABCDEFGH", [24, 22, 48, 18, 30, 30, 8, 12]):
        ws.column_dimensions[col].width = w
    make_banner(ws,
        f"DIVERGÊNCIAS NA EXTRAÇÃO — {len(MODELOS)} MODELOS",
        f"Campos com valor extraído diferente do esperado — {len(TIPOS)} tipos de documento",
        8)

    ws.row_dimensions[4].height = 24
    for ci, h in enumerate(
        ["Tipo", "Modelo", "Arquivo", "Campo", "Esperado", "Extraído", "Match", "Levenshtein"], 1
    ):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(bold=True, size=9, color="FFFFFF")
        c.fill = fill(BASE_COLORS["err_hdr"])
        c.alignment = aln(wrap=True); c.border = bdr()

    row3 = 5; div_count = 0
    for tipo in TIPOS:
        for model in MODELOS:
            d = all_data.get((tipo, model))
            if not d: continue
            for arq in d.get("resultados_por_arquivo", []):
                errados = [c for c in arq.get("campos", []) if not c.get("match", True)]
                if not errados: continue
                arquivo = arq["arquivo"]
                tipo_label = tipo.upper().replace("_", " ")
                model_label = MODEL_LABELS.get(model, model)
                for cd in errados:
                    div_count += 1
                    bg = BASE_COLORS["err_light"] if row3 % 2 == 0 else BASE_COLORS["err_alt"]
                    ws.row_dimensions[row3].height = 20
                    vals = [
                        tipo_label, model_label, arquivo,
                        cd.get("campo", ""), str(cd.get("esperado", "")),
                        str(cd.get("extraido", "")), "✗", cd.get("levenshtein"),
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=row3, column=ci)
                        if isinstance(v, float):
                            c.value = round(v, 4); c.number_format = "0.0000"
                        else:
                            c.value = v if v is not None else "—"
                        c.fill = fill(bg)
                        c.alignment = (aln(h="left", wrap=True) if ci >= 3
                                       else aln(h="left"))
                        c.font = fnt(size=9, color=("B71C1C" if ci == 7 else "000000"))
                        c.border = bdr()
                    row3 += 1

    ws.row_dimensions[row3].height = 18
    ws.merge_cells(f"A{row3}:H{row3}")
    c = ws.cell(row=row3, column=1, value=f"Total de divergências: {div_count}")
    c.font = fnt(bold=True, size=10, color="FFFFFF")
    c.fill = fill(BASE_COLORS["err_hdr"]); c.alignment = aln(h="left")
    for col in range(2, 9):
        ws.cell(row=row3, column=col).fill = fill(BASE_COLORS["err_hdr"])
    return div_count


# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — Todos os Arquivos
# ══════════════════════════════════════════════════════════════════════════════
def build_todos_arquivos(wb, all_data, TIPOS, MODELOS, MODEL_LABELS):
    ws = wb.create_sheet("📄 Todos os Arquivos")
    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 48
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12
    make_banner(ws,
        "RESULTADOS INDIVIDUAIS POR ARQUIVO",
        f"Métricas de extração por arquivo processado — {len(TIPOS)} tipos · {len(MODELOS)} modelos",
        10)

    ws.row_dimensions[4].height = 24
    for ci, h in enumerate(
        ["Tipo","Modelo","Arquivo","Accuracy","F1","Precisão",
         "Recall","Levenshtein","Latência (s)","Divergências"], 1
    ):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(bold=True, size=9, color="FFFFFF")
        c.fill = fill(BASE_COLORS["title_bg"])
        c.alignment = aln(wrap=True); c.border = bdr()

    row4 = 5
    for tipo in TIPOS:
        for model in MODELOS:
            d = all_data.get((tipo, model))
            if not d: continue
            for arq in d.get("resultados_por_arquivo", []):
                bg = BASE_COLORS["alt1"] if row4 % 2 == 0 else BASE_COLORS["alt2"]
                ws.row_dimensions[row4].height = 15
                m = arq.get("metricas", {})
                erros = sum(1 for c in arq.get("campos", [])
                            if not c.get("match", True))
                vals = [
                    tipo.upper().replace("_", " "),
                    MODEL_LABELS.get(model, model),
                    arq["arquivo"],
                    m.get("accuracy"), m.get("f1"),
                    m.get("precision"), m.get("recall"),
                    m.get("levenshtein"),
                    round(m.get("latencia", 0), 2) if m.get("latencia") else "—",
                    f"{erros} campo(s)" if erros else "—",
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row4, column=ci)
                    if isinstance(v, float):
                        cell.value = round(v, 4)
                        cell.number_format = "0.00%" if ci <= 8 else "0.00"
                    else:
                        cell.value = v
                    cell.fill = fill(bg)
                    cell.alignment = aln(h="left") if ci == 3 else aln()
                    cell.font = fnt(size=9,
                                    color=("B71C1C" if ci == 10 and erros else "000000"))
                    cell.border = bdr()
                row4 += 1

    for cl in ["D", "E", "F", "G", "H"]:
        ws.conditional_formatting.add(f"{cl}5:{cl}{row4}", cscale())
    return row4 - 5


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    # Argumentos
    root_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("fce-modulo5/results")
    out_file = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("fce-modulo5/resultados_extracao.xlsx")

    if not root_dir.exists():
        # Tenta modo relativo se falhar
        if Path("resultados").exists():
            root_dir = Path("resultados")
        else:
            print(f"[ERRO] Pasta não encontrada: {root_dir}")
            sys.exit(1)

    print(f"📂 Lendo JSONs em: {root_dir}")
    all_data = load_data(root_dir)

    if not all_data:
        print("[ERRO] Nenhum arquivo JSON válido encontrado.")
        sys.exit(1)

    TIPOS   = sorted(set(k[0] for k in all_data))
    MODELOS_BRUTO = sorted(set(k[1] for k in all_data))

    # Calcula média de acurácia por modelo para ordenação
    model_avg_acc = {}
    for m in MODELOS_BRUTO:
        accs = [all_data[(t, m)]["metricas_gerais"].get("mean_accuracy", 0) 
                for t in TIPOS if (t, m) in all_data]
        model_avg_acc[m] = sum(accs) / len(accs) if accs else 0
    
    # Ordena modelos pela acurácia média (decrescente)
    MODELOS = sorted(MODELOS_BRUTO, key=lambda m: model_avg_acc[m], reverse=True)

    MODEL_LABELS, MODEL_COLORS = build_model_meta(MODELOS)

    # Extrair intervalo de datas dos timestamps
    timestamps = [d["timestamp"][:10] for d in all_data.values()]
    date_min = min(timestamps); date_max = max(timestamps)
    date_range = date_min if date_min == date_max else f"{date_min} / {date_max}"

    print(f"✅ {len(all_data)} registros carregados")
    print(f"   Tipos     : {len(TIPOS)} → {TIPOS}")
    print(f"   Modelos   : {len(MODELOS)} → {MODELOS}")
    print(f"   Período   : {date_range}")
    print()

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    print("📊 Gerando aba: Visão Geral...")
    build_visao_geral(wb, all_data, TIPOS, MODELOS, MODEL_LABELS, MODEL_COLORS, date_range)

    print("🔍 Gerando aba: Métricas por Campo...")
    build_metricas_campo(wb, all_data, TIPOS, MODELOS, MODEL_LABELS, MODEL_COLORS)

    print("⚠️  Gerando aba: Divergências...")
    div_count = build_divergencias(wb, all_data, TIPOS, MODELOS, MODEL_LABELS)

    print("📄 Gerando aba: Todos os Arquivos...")
    arq_count = build_todos_arquivos(wb, all_data, TIPOS, MODELOS, MODEL_LABELS)

    wb.save(out_file)
    print()
    print(f"✅ Planilha salva em: {out_file}")
    print(f"   Divergências : {div_count}")
    print(f"   Arquivos     : {arq_count}")


if __name__ == "__main__":
    main()