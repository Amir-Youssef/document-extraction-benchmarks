import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuração ---
V1_DIR = Path("fce-modulo5/results_v1")
V2_DIR = Path("fce-modulo5/results")
OUTPUT_EXCEL = Path("fce-modulo5/comparativo_v1_vs_v2.xlsx")

# Mapeamento para normalizar nomes de modelos diferentes entre as versões
MODEL_MAP = {
    "gemini-pro": "gemini31_pro",
    "gemini-flash": "gemini_flash",
    "gemini31_pro": "gemini31_pro",
    "gemini_flash": "gemini_flash",
    # Adicione outros mapeamentos se necessário
}

def normalize_model(m):
    return MODEL_MAP.get(m, m).lower().replace("-", "_")

def load_all_jsons(directory):
    data = {}
    if not directory.exists():
        print(f"⚠️  Diretório não encontrado: {directory}")
        return data
        
    for f in directory.rglob("*.json"):
        try:
            with open(f, encoding="utf-8") as fp:
                d = json.load(fp)
            tipo = d.get("tipo_documento", "")
            model = normalize_model(d.get("modelo", ""))
            key = (tipo, model)
            
            # Se houver duplicatas por causa de timestamps, pegar o mais recente
            if key not in data or d.get("timestamp", "") > data[key].get("timestamp", ""):
                data[key] = d
        except Exception as e:
            print(f"Erro ao ler {f}: {e}")
    return data

def build_comparison():
    print(f"📂 Lendo V1 em: {V1_DIR}")
    data_v1 = load_all_jsons(V1_DIR)
    print(f"📂 Lendo V2 em: {V2_DIR}")
    data_v2 = load_all_jsons(V2_DIR)

    common_keys = sorted(set(data_v1.keys()) & set(data_v2.keys()))
    
    if not common_keys:
        print("⚠️  Nenhum par (Tipo, Modelo) em comum encontrado para comparar.")
        # Mostrar o que foi encontrado em cada uma para ajudar o usuário
        print(f"V1: {len(data_v1)} itens | V2: {len(data_v2)} itens")
        print(f"Exemplos V1: {list(data_v1.keys())[:3]}")
        print(f"Exemplos V2: {list(data_v2.keys())[:3]}")
        return []

    print(f"✅ Encontrados {len(common_keys)} itens para comparação.")
    
    results = []
    for key in common_keys:
        v1 = data_v1[key].get("metricas_gerais", {})
        v2 = data_v2[key].get("metricas_gerais", {})
        
        res = {
            "tipo": key[0],
            "modelo": key[1],
            "acc_v1": v1.get("mean_accuracy", 0),
            "acc_v2": v2.get("mean_accuracy", 0),
            "lev_v1": v1.get("mean_levenshtein", 0),
            "lev_v2": v2.get("mean_levenshtein", 0),
            "lat_v1": v1.get("latencia", {}).get("mean", 0),
            "lat_v2": v2.get("latencia", {}).get("mean", 0),
        }
        # Ganho/Perda
        res["acc_diff"] = res["acc_v2"] - res["acc_v1"]
        res["lev_diff"] = res["lev_v2"] - res["lev_v1"]
        res["lat_diff"] = res["lat_v2"] - res["lat_v1"]
        
        results.append(res)
        
    return results

def save_to_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo V1 vs V2"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    center = Alignment(horizontal="center")
    
    headers = [
        "Tipo", "Modelo", 
        "Accuracy V1", "Accuracy V2", "Δ Acc",
        "Levenshtein V1", "Levenshtein V2", "Δ Lev",
        "Latência V1", "Latência V2", "Δ Lat"
    ]
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = 20

    for r, res in enumerate(results, 2):
        ws.cell(row=r, column=1, value=res["tipo"])
        ws.cell(row=r, column=2, value=res["modelo"])
        
        # Accuracy
        ws.cell(row=r, column=3, value=res["acc_v1"]).number_format = "0.00%"
        ws.cell(row=r, column=4, value=res["acc_v2"]).number_format = "0.00%"
        diff_acc = ws.cell(row=r, column=5, value=res["acc_diff"])
        diff_acc.number_format = "+0.00%;-0.00%;0.00%"
        
        # Levenshtein
        ws.cell(row=r, column=6, value=res["lev_v1"]).number_format = "0.00%"
        ws.cell(row=r, column=7, value=res["lev_v2"]).number_format = "0.00%"
        diff_lev = ws.cell(row=r, column=8, value=res["lev_diff"])
        diff_lev.number_format = "+0.00%;-0.00%;0.00%"
        
        # Latência
        ws.cell(row=r, column=9, value=res["lat_v1"]).number_format = "0.00s"
        ws.cell(row=r, column=10, value=res["lat_v2"]).number_format = "0.00s"
        diff_lat = ws.cell(row=r, column=11, value=res["lat_diff"])
        diff_lat.number_format = "+0.00s;-0.00s;0.00s"
        
        # Cores para Delta
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        if abs(res["acc_diff"]) > 0.0001:
            if res["acc_diff"] > 0: diff_acc.fill = green
            else: diff_acc.fill = red
        
        if abs(res["lev_diff"]) > 0.0001:
            if res["lev_diff"] > 0: diff_lev.fill = green
            else: diff_lev.fill = red
        
        if abs(res["lat_diff"]) > 0.1:
            if res["lat_diff"] < 0: diff_lat.fill = green # Latência menor é melhor
            else: diff_lat.fill = red

    wb.save(OUTPUT_EXCEL)
    print(f"✅ Planilha salva em: {OUTPUT_EXCEL}")

def main():
    results = build_comparison()
    if results:
        save_to_excel(results)
        print("\n--- RESUMO DE IMPACTO (Geral) ---")
        avg_acc = sum(r["acc_diff"] for r in results) / len(results)
        print(f"Variação média de Accuracy: {avg_acc:+.2%}")
        
if __name__ == "__main__":
    main()
